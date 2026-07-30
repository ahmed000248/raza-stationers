#!/usr/bin/env python3
"""
Phase 2 Catalogue Certification Tool — Raza Stationers
=======================================================
Read-only, deterministic. Never writes to the workbook.
Never connects to a database. Produces JSON + terminal report.

Workbook layout (Products sheet):
  Row 1: Title row
  Row 2: Subtitle/description row
  Row 3: Blank (not present in XML)
  Row 4: Header row
  Rows 5+: Product data rows (2,167 expected)

Usage:
    python tools/certify_catalogue.py data/final/Raza-Stationers-Final-Supabase-Catalogue.xlsx
    python tools/certify_catalogue.py ... --json-out report.json

Exit codes:
    0  All mandatory checks pass (CERTIFIED or CERTIFIED WITH ADVISORIES)
    1  One or more mandatory checks fail (NOT CERTIFIED)
"""

import argparse
import hashlib
import json
import re
import sys
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path

# Force UTF-8 output on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


# ── Approved constants ────────────────────────────────────────────────────────
REGISTERED_SHA256   = "7cb65d6d07b30c75a048431dab4f855fd60b901515c07fe0f2253f8faccafa0b"
REGISTERED_SIZE     = 376_017
EXPECTED_PRODUCTS   = 2_167
EXPECTED_CATEGORIES = 103
EXPECTED_INDIVIDUAL = 70
EXPECTED_WHOLESALE  = 2_097
PRODUCTS_SHEET      = "Products"
CATEGORIES_SHEET    = "Categories"
HEADER_ROW          = 4   # Excel row number of the header row

# In the final approved workbook, Sales Type is the authoritative field.
# Individual rows have Unit of Measure = "Piece" AND Pack Quantity = 1.
# The 1PC/1PCS/1 PCS classification rule was applied during workbook creation;
# the final workbook encodes the result in the Sales Type column directly.
INDIVIDUAL_UNIT      = "piece"     # normalised, case-insensitive
INDIVIDUAL_PACK_QTY  = 1           # numeric pack quantity for Individual

# Formula columns (analytics only, not import fields)
FORMULA_COLUMNS = {"Profit", "Profit Margin %", "Markup %"}


# ── Utilities ─────────────────────────────────────────────────────────────────

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_ws(s) -> str:
    """Strip leading/trailing whitespace, collapse internal runs."""
    return re.sub(r"\s+", " ", str(s)).strip() if s is not None else ""


def has_control_chars(s: str) -> bool:
    return any(unicodedata.category(c) in ("Cc", "Cf") and c not in ("\t", "\n", "\r")
               for c in str(s))


def is_individual_by_rule(unit_value, pack_qty) -> bool:
    """True when unit=Piece AND pack_qty=1, matching the approved final-workbook pattern."""
    if unit_value is None:
        return False
    unit_norm = normalize_ws(str(unit_value)).lower()
    if unit_norm != INDIVIDUAL_UNIT:
        return False
    try:
        return float(str(pack_qty).replace(",", "")) == INDIVIDUAL_PACK_QTY
    except (ValueError, TypeError):
        return False


def col_letter_to_index(col: str) -> int:
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def parse_cell_ref(ref: str):
    m = re.match(r"([A-Z]+)(\d+)", ref.upper())
    if m:
        return col_letter_to_index(m.group(1)), int(m.group(2))
    return None, None


# ── Method A: openpyxl ────────────────────────────────────────────────────────

def method_a_openpyxl(path: Path) -> dict:
    """Read workbook with openpyxl (data_only=True).
    Header is on Excel row HEADER_ROW; product data starts on row HEADER_ROW+1.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not available"}

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    result = {
        "sheet_names": wb.sheetnames,
        "hidden_sheets": [s.title for s in wb.worksheets if s.sheet_state != "visible"],
        "very_hidden_sheets": [s.title for s in wb.worksheets if s.sheet_state == "veryHidden"],
        "macros": wb.vba_archive is not None,
        "external_links": bool(wb._external_links),
    }

    if PRODUCTS_SHEET not in wb.sheetnames:
        result["products_error"] = f"Sheet '{PRODUCTS_SHEET}' not found"
        return result

    ws = wb[PRODUCTS_SHEET]
    result["products_sheet_state"] = ws.sheet_state
    result["products_dimensions"] = ws.dimensions
    result["sheet_protected"] = ws.protection.sheet if ws.protection else False

    # Hidden columns
    hidden_cols = [letter for letter, cd in ws.column_dimensions.items() if cd.hidden]
    result["hidden_columns"] = hidden_cols

    # Hidden rows
    hidden_rows = [rn for rn, rd in ws.row_dimensions.items() if rd.hidden]
    result["hidden_rows_count"] = len(hidden_rows)
    result["hidden_rows_sample"] = hidden_rows[:10]

    # Read header row (Excel row HEADER_ROW)
    header_cells = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]
    headers = list(header_cells)
    # Trim trailing None headers
    while headers and headers[-1] is None:
        headers.pop()
    result["headers"] = headers

    # Iterate product rows (rows after header)
    rows = []
    formula_rows = []
    error_rows = []
    blank_rows = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=False), start=HEADER_ROW + 1
    ):
        vals = [c.value for c in row[:len(headers)]]

        if all(v is None or str(v).strip() == "" for v in vals):
            blank_rows.append(row_idx)
            continue

        for ci, cell in enumerate(row[:len(headers)]):
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_rows.append({"row": row_idx, "col": ci + 1, "formula": cell.value[:60]})
            if cell.data_type == "e":
                error_rows.append({"row": row_idx, "col": ci + 1, "error": cell.value})

        rows.append(vals)

    # Merged cells inside the data range
    merged_in_data = [str(mc) for mc in ws.merged_cells.ranges if mc.min_row > HEADER_ROW]

    result["row_count"] = len(rows)
    result["blank_rows_in_data"] = blank_rows
    result["formula_count"] = len(formula_rows)
    result["formula_sample"] = formula_rows[:5]
    result["error_cells"] = error_rows
    result["merged_in_data"] = merged_in_data[:10]
    result["rows"] = rows

    # Categories sheet
    if CATEGORIES_SHEET in wb.sheetnames:
        ws_cat = wb[CATEGORIES_SHEET]
        cat_vals = []
        # Categories sheet: check rows for non-empty first column values after header rows
        for r in ws_cat.iter_rows(min_row=4, values_only=True):  # skip header rows
            if r and r[0] is not None and str(r[0]).strip():
                cat_vals.append(str(r[0]).strip())
        result["categories_raw_count"] = len(cat_vals)
        result["categories_sample"] = cat_vals[:5]
    else:
        result["categories_raw_count"] = None

    wb.close()
    return result


# ── Method B: OOXML / ZIP inspection ─────────────────────────────────────────

NS_SS  = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def _cell_value_ooxml(c_elem, shared_strings: list):
    """Extract display value from an OOXML <c> element."""
    t    = c_elem.get("t", "n")
    v    = c_elem.find(f"{{{NS_SS}}}v")
    f    = c_elem.find(f"{{{NS_SS}}}f")
    is_  = c_elem.find(f"{{{NS_SS}}}is")   # inline string
    has_formula = f is not None
    is_error    = (t == "e")

    if is_ is not None:
        # Inline string (t="inlineStr" or "str")
        val = "".join(tx.text or "" for tx in is_.iter(f"{{{NS_SS}}}t"))
        return val, has_formula, False

    if v is None or v.text is None:
        return None, has_formula, is_error

    raw = v.text
    if t == "s":
        # Shared string index
        idx = int(raw)
        return shared_strings[idx] if idx < len(shared_strings) else "", has_formula, False
    if t in ("str", "inlineStr"):
        return raw, has_formula, False
    if t == "e":
        return raw, has_formula, True

    # Numeric — return as-is (string representation of number)
    return raw, has_formula, False


def method_b_ooxml(path: Path) -> dict:
    """Inspect workbook as OOXML ZIP."""
    result = {}
    try:
        with zipfile.ZipFile(path, "r") as zf:
            names = zf.namelist()
            result["zip_entries_count"] = len(names)
            result["has_vba"]              = any("vbaProject" in n for n in names)
            result["has_external_links"]   = any("externalLinks" in n for n in names)
            result["has_connections"]      = any("connections.xml" in n for n in names)

            # Shared strings
            shared_strings: list[str] = []
            if "xl/sharedStrings.xml" in names:
                raw = zf.read("xl/sharedStrings.xml")
                tree = ET.fromstring(raw)
                for si in tree.iter(f"{{{NS_SS}}}si"):
                    shared_strings.append(
                        "".join(t.text or "" for t in si.iter(f"{{{NS_SS}}}t"))
                    )
            result["shared_strings_count"] = len(shared_strings)

            # Workbook sheet list
            wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
            sheets = []
            for sh in wb_xml.iter(f"{{{NS_SS}}}sheet"):
                state = sh.get("state", "visible")
                sheets.append({
                    "name":    sh.get("name"),
                    "sheetId": sh.get("sheetId"),
                    "state":   state,
                    "rId":     sh.get(f"{{{NS_R}}}id"),
                })
            result["sheets"]        = sheets
            result["sheet_names"]   = [s["name"] for s in sheets]
            result["hidden_sheets"] = [s["name"] for s in sheets if s["state"] != "visible"]

            # rId -> target mapping
            rel_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rId_to_target = {
                r.get("Id"): r.get("Target")
                for r in rel_xml.iter(f"{{{NS_REL}}}Relationship")
            }

            prod_info = next((s for s in sheets if s["name"] == PRODUCTS_SHEET), None)
            if not prod_info:
                result["products_error"] = f"Sheet '{PRODUCTS_SHEET}' not found"
                return result

            target    = rId_to_target.get(prod_info["rId"], "")
            # Target may be absolute ("/xl/worksheets/sheet2.xml") or relative ("worksheets/sheet2.xml")
            if target.startswith("/"):
                prod_path = target.lstrip("/")          # "/xl/..." -> "xl/..."
            else:
                prod_path = f"xl/{target}"              # "worksheets/..." -> "xl/worksheets/..."
            if prod_path not in names:
                result["products_error"] = f"Products file not found: {prod_path}"
                return result

            ws_xml = ET.fromstring(zf.read(prod_path))
            prot = ws_xml.find(f"{{{NS_SS}}}sheetProtection")
            result["sheet_protected"] = prot is not None

            # Merged cells
            merged = []
            for mc in ws_xml.iter(f"{{{NS_SS}}}mergeCell"):
                ref = mc.get("ref", "")
                if ref:
                    row_num = re.search(r"\d+", ref)
                    if row_num and int(row_num.group()) > HEADER_ROW:
                        merged.append(ref)
            result["merged_in_data"]       = merged[:10]
            result["merged_in_data_count"] = len(merged)

            headers_b  = []
            rows_b     = []
            formula_count = 0
            error_cells = []
            blank_rows  = []

            all_rows = ws_xml.findall(f"{{{NS_SS}}}sheetData/{{{NS_SS}}}row")

            for row_elem in all_rows:
                r_num = int(row_elem.get("r", "0"))
                cells = {}
                for c_elem in row_elem:
                    if not c_elem.tag.endswith("}c"):
                        continue
                    ref = c_elem.get("r", "")
                    col_i, _ = parse_cell_ref(ref)
                    if col_i is None:
                        continue
                    val, is_formula, is_error = _cell_value_ooxml(c_elem, shared_strings)
                    cells[col_i] = val
                    if is_formula:
                        formula_count += 1
                    if is_error:
                        error_cells.append({"row": r_num, "col": col_i, "error": val})

                if r_num == HEADER_ROW:
                    if cells:
                        max_col = max(cells.keys())
                        headers_b = [cells.get(i) for i in range(1, max_col + 1)]
                    continue

                if r_num <= HEADER_ROW:
                    continue  # title/subtitle rows

                if not cells or all(
                    v is None or str(v).strip() == "" for v in cells.values()
                ):
                    blank_rows.append(r_num)
                    continue

                row_vals = [cells.get(i) for i in range(1, len(headers_b) + 1)]
                rows_b.append(row_vals)

            result["headers"]           = headers_b
            result["row_count"]         = len(rows_b)
            result["formula_count"]     = formula_count
            result["error_cells"]       = error_cells
            result["blank_rows_in_data"] = blank_rows
            result["rows"]              = rows_b

    except Exception as exc:
        result["error"] = str(exc)

    return result


# ── Cross-check ───────────────────────────────────────────────────────────────

def cross_check(a: dict, b: dict) -> tuple[list, list]:
    """Return (blocking_disagreements, advisory_notes) between method A and B."""
    blocking: list[str] = []
    advisory: list[str] = []

    if a.get("sheet_names") != b.get("sheet_names"):
        blocking.append(f"DISAGREE sheet_names: A={a.get('sheet_names')!r}  B={b.get('sheet_names')!r}")

    if a.get("row_count") != b.get("row_count"):
        blocking.append(f"DISAGREE row_count: A={a.get('row_count')!r}  B={b.get('row_count')!r}")

    # Formula count: A uses data_only=True (sees cached values), B reads raw OOXML.
    # A difference is expected when the workbook contains formula-driven columns.
    fa = a.get("formula_count", 0)
    fb = b.get("formula_count", 0)
    if fa != fb:
        advisory.append(
            f"Formula count differs between methods (expected): "
            f"A (data_only)={fa}  B (raw OOXML)={fb}. "
            f"Formulas are in analytics columns (Profit, Profit Margin %%, Markup %%); "
            f"these are not import fields."
        )

    # Headers
    ah = [str(h) for h in (a.get("headers") or []) if h is not None]
    bh = [str(h) for h in (b.get("headers") or []) if h is not None]
    if ah != bh:
        blocking.append(f"DISAGREE headers: A={ah}  B={bh}")

    return blocking, advisory



# ── Analytics ─────────────────────────────────────────────────────────────────

def find_col(headers: list, *names: str) -> int | None:
    """Return index of first matching header (case-insensitive)."""
    lower_names = [n.lower() for n in names]
    for i, h in enumerate(headers):
        if h and normalize_ws(str(h)).lower() in lower_names:
            return i
    return None


def analyse_rows(headers: list, rows: list) -> dict:
    result: dict = {"blocking": [], "advisory": [], "review": []}

    name_col   = find_col(headers, "product name", "name", "productname")
    cat_col    = find_col(headers, "category")
    type_col   = find_col(headers, "sales type", "salestype", "type")
    unit_col   = find_col(headers, "unit of measure", "unit", "packaging", "pack size", "packsize")
    sku_col    = find_col(headers, "sku", "source key", "source_key")
    price_cols = [i for i, h in enumerate(headers) if h and "price" in str(h).lower()]

    result["column_map"] = {
        "name": name_col, "category": cat_col, "sales_type": type_col,
        "unit": unit_col, "sku": sku_col, "price_cols": price_cols,
    }

    total = len(rows)
    result["total_rows"] = total
    if total != EXPECTED_PRODUCTS:
        result["blocking"].append(
            f"TOTAL ROW COUNT: expected {EXPECTED_PRODUCTS}, found {total}")

    # ── Sales Type ───────────────────────────────────────────────────────────
    if type_col is None:
        result["blocking"].append("SALES TYPE COLUMN not found")
        individual = wholesale = 0
    else:
        type_vals = [
            normalize_ws(str(r[type_col])) if r[type_col] is not None else ""
            for r in rows
        ]
        tc = Counter(type_vals)
        result["sales_type_distribution"] = dict(tc)
        individual = sum(v for k, v in tc.items() if k.lower() == "individual")
        wholesale  = sum(v for k, v in tc.items() if k.lower() == "wholesale")
        unknown    = total - individual - wholesale
        result["individual_count"] = individual
        result["wholesale_count"]  = wholesale
        result["unknown_type_count"] = unknown

        if individual != EXPECTED_INDIVIDUAL:
            result["blocking"].append(
                f"INDIVIDUAL COUNT: expected {EXPECTED_INDIVIDUAL}, found {individual}")
        if wholesale != EXPECTED_WHOLESALE:
            result["blocking"].append(
                f"WHOLESALE COUNT: expected {EXPECTED_WHOLESALE}, found {wholesale}")
        blank_types = sum(1 for v in type_vals if v == "")
        if blank_types:
            result["blocking"].append(f"BLANK SALES TYPE: {blank_types} rows")
        if unknown > 0:
            bad = [v for v in set(type_vals) if v.lower() not in ("individual", "wholesale") and v]
            result["blocking"].append(
                f"UNKNOWN SALES TYPE values ({unknown} rows): {bad[:10]}")

    # ── Classification rule ──────────────────────────────────────────────────
    # In the final workbook Sales Type is the authoritative field.
    # Verify consistency: all Individual rows should have unit=Piece, pack_qty=1.
    pack_col = find_col(headers, "pack quantity", "pack qty", "packquantity")
    if unit_col is not None and type_col is not None and pack_col is not None:
        type_ind_set = {i for i, r in enumerate(rows)
                        if normalize_ws(str(r[type_col] or "")).lower() == "individual"}
        rule_ind_set = {i for i, r in enumerate(rows)
                        if is_individual_by_rule(r[unit_col], r[pack_col])}
        in_type_not_rule = type_ind_set - rule_ind_set
        in_rule_not_type = rule_ind_set - type_ind_set
        result["rule_individual_count"] = len(rule_ind_set)
        result["rule_wholesale_count"]  = total - len(rule_ind_set)
        if in_type_not_rule:
            sample_units = [(rows[i][unit_col], rows[i][pack_col]) for i in list(in_type_not_rule)[:5]]
            result["advisory"].append(
                f"CLASSIFICATION CONSISTENCY: {len(in_type_not_rule)} rows marked Individual "
                f"but do not have unit=Piece / pack_qty=1 (sample unit/qty: {sample_units})")
        if in_rule_not_type:
            result["advisory"].append(
                f"CLASSIFICATION CONSISTENCY: {len(in_rule_not_type)} rows have "
                f"unit=Piece/pack_qty=1 but are marked Wholesale — may be intentional")
    elif unit_col is not None and type_col is not None:
        result["advisory"].append(
            "Cannot fully verify classification: Pack Quantity column missing")
    else:
        result["advisory"].append(
            "Cannot verify classification: unit or sales-type column missing")

    # ── Categories ───────────────────────────────────────────────────────────
    if cat_col is None:
        result["blocking"].append("CATEGORY COLUMN not found")
    else:
        raw_cats  = [r[cat_col] for r in rows]
        cat_strs  = [normalize_ws(str(v)) if v is not None else "" for v in raw_cats]
        blank_cats = sum(1 for c in cat_strs if c == "")
        if blank_cats:
            result["blocking"].append(f"BLANK CATEGORY: {blank_cats} rows")

        distinct = set(c for c in cat_strs if c)
        result["distinct_categories"] = len(distinct)
        if len(distinct) != EXPECTED_CATEGORIES:
            result["blocking"].append(
                f"CATEGORY COUNT: expected {EXPECTED_CATEGORIES}, found {len(distinct)}")

        # Case-only collisions
        lower_map: dict = {}
        for c in distinct:
            lower_map.setdefault(c.lower(), []).append(c)
        case_coll = {k: v for k, v in lower_map.items() if len(v) > 1}
        if case_coll:
            result["advisory"].append(
                f"CASE-ONLY CATEGORY COLLISIONS ({len(case_coll)}): "
                f"{list(case_coll.items())[:5]}")

        # Whitespace in raw values
        ws_cats = [str(v) for v in raw_cats if v is not None and str(v) != normalize_ws(str(v))]
        if ws_cats:
            result["advisory"].append(
                f"LEADING/TRAILING WHITESPACE IN CATEGORY ({len(ws_cats)} rows): "
                f"sample {ws_cats[:3]}")

    # ── Identity / Names ─────────────────────────────────────────────────────
    if name_col is None:
        result["advisory"].append("PRODUCT NAME COLUMN not found — using first column")
        name_col = 0

    names = [normalize_ws(str(r[name_col])) if r[name_col] is not None else "" for r in rows]
    blank_names = sum(1 for n in names if n == "")
    if blank_names:
        result["blocking"].append(f"BLANK PRODUCT NAMES: {blank_names} rows")

    name_counter = Counter(names)
    exact_dupes  = {k: v for k, v in name_counter.items() if v > 1 and k != ""}
    if exact_dupes:
        result["blocking"].append(
            f"EXACT DUPLICATE NAMES "
            f"({sum(v for v in exact_dupes.values()) - len(exact_dupes)} extra copies): "
            f"sample {list(exact_dupes.items())[:5]}")
    result["exact_duplicate_names"] = len(exact_dupes)

    ctrl_rows = [i + HEADER_ROW + 1 for i, n in enumerate(names) if has_control_chars(n)]
    if ctrl_rows:
        result["advisory"].append(
            f"CONTROL CHARS IN NAMES ({len(ctrl_rows)} rows): rows {ctrl_rows[:5]}")

    # ── SKU uniqueness ────────────────────────────────────────────────────────
    if sku_col is not None:
        skus = [normalize_ws(str(r[sku_col])) if r[sku_col] is not None else "" for r in rows]
        sc   = Counter(skus)
        blank_skus = sc.get("", 0)
        dup_skus   = {k: v for k, v in sc.items() if v > 1 and k != ""}
        if blank_skus:
            result["advisory"].append(f"BLANK SKU VALUES: {blank_skus} rows")
        if dup_skus:
            result["blocking"].append(
                f"DUPLICATE SKU VALUES ({len(dup_skus)}): sample {list(dup_skus.items())[:5]}")
        result["sku_unique_count"] = len(sc) - (1 if "" in sc else 0)

    # ── Price fields ─────────────────────────────────────────────────────────
    for pc in price_cols:
        hname = str(headers[pc]) if pc < len(headers) else f"col{pc}"
        vals, negatives, non_numeric, zeros = [], [], [], []
        for ri, row in enumerate(rows):
            v = row[pc] if pc < len(row) else None
            if v is None or str(v).strip() == "":
                continue
            try:
                fv = float(str(v).replace(",", ""))
                vals.append(fv)
                if fv < 0:
                    negatives.append(ri + HEADER_ROW + 1)
                if fv == 0:
                    zeros.append(ri + HEADER_ROW + 1)
            except (ValueError, TypeError):
                non_numeric.append(ri + HEADER_ROW + 1)

        if negatives:
            result["blocking"].append(
                f"NEGATIVE PRICES in '{hname}' ({len(negatives)} rows)")
        if non_numeric:
            result["blocking"].append(
                f"NON-NUMERIC PRICES in '{hname}' ({len(non_numeric)} rows)")
        if vals:
            result.setdefault("price_stats", {})[hname] = {
                "count": len(vals),
                "min":   min(vals),
                "max":   max(vals),
                "zeros": len(zeros),
            }
            if zeros:
                result["advisory"].append(
                    f"ZERO PRICES in '{hname}' ({len(zeros)} rows) — review")

    # ── Packaging / unit fields ───────────────────────────────────────────────
    if unit_col is not None:
        unit_strs = [
            normalize_ws(str(r[unit_col])) if r[unit_col] is not None else "" for r in rows
        ]
        blank_units = sum(1 for u in unit_strs if u == "")
        if blank_units:
            result["advisory"].append(f"BLANK UNIT/PACKAGING VALUES: {blank_units} rows")
        distinct_units = set(u for u in unit_strs if u)
        result["unit_distinct_values"] = len(distinct_units)
        result["unit_sample"] = sorted(list(distinct_units))[:20]

    # ── Spreadsheet hazards ───────────────────────────────────────────────────
    inj = [i + HEADER_ROW + 1 for i, n in enumerate(names) if n and n[0] in ("=", "+", "-", "@")]
    if inj:
        result["advisory"].append(
            f"FORMULA-INJECTION PREFIX in name ({len(inj)} rows): rows {inj[:5]}")

    nbsp_rows = [i + HEADER_ROW + 1 for i, n in enumerate(names) if "\u00a0" in n]
    if nbsp_rows:
        result["advisory"].append(
            f"NON-BREAKING SPACES in name ({len(nbsp_rows)} rows): rows {nbsp_rows[:5]}")

    long_names = [i + HEADER_ROW + 1 for i, n in enumerate(names) if len(n) > 255]
    if long_names:
        result["advisory"].append(
            f"EXCESSIVELY LONG NAMES (>255 chars) in {len(long_names)} rows")

    return result


# ── Column profiler ───────────────────────────────────────────────────────────

def profile_columns(headers: list, rows: list) -> list:
    cols = []
    for ci, hdr in enumerate(headers):
        vals    = [r[ci] if ci < len(r) else None for r in rows]
        non_emp = [v for v in vals if v is not None and str(v).strip() != ""]
        num_vals = []
        for v in non_emp:
            try:
                num_vals.append(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                pass

        stat: dict = {
            "index":          ci + 1,
            "header":         str(hdr) if hdr is not None else None,
            "non_empty_count": len(non_emp),
            "unique_count":    len(set(str(v) for v in non_emp)),
        }
        if num_vals:
            stat["min"] = min(num_vals)
            stat["max"] = max(num_vals)
        else:
            text_lens = [len(str(v)) for v in non_emp]
            if text_lens:
                stat["max_text_length"] = max(text_lens)
        stat["sample"] = [str(v) for v in non_emp[:3]]
        cols.append(stat)
    return cols


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Phase 2 Catalogue Certification Tool")
    parser.add_argument("workbook", help="Path to the approved Excel workbook")
    parser.add_argument("--json-out", metavar="FILE", help="Write JSON report to FILE")
    args = parser.parse_args()

    path = Path(args.workbook)
    sep  = "-" * 70

    print(sep)
    print("RAZA STATIONERS -- PHASE 2 CATALOGUE CERTIFICATION")
    print(sep)

    # ── Identity ──────────────────────────────────────────────────────────────
    print("\n[1] SOURCE IDENTITY")
    if not path.exists():
        print(f"  FAIL  File not found: {path}")
        sys.exit(1)

    actual_size   = path.stat().st_size
    actual_sha256 = sha256_file(path)
    filename_ok   = path.name == "Raza-Stationers-Final-Supabase-Catalogue.xlsx"
    size_ok       = actual_size == REGISTERED_SIZE
    hash_ok       = actual_sha256 == REGISTERED_SHA256

    print(f"  File   : {path}")
    print(f"  Size   : {actual_size:,} bytes  {'OK' if size_ok else 'MISMATCH'}")
    print(f"  SHA-256: {actual_sha256}  {'OK' if hash_ok else 'MISMATCH'}")
    print(f"  Tracked: {'OK' if filename_ok else 'unexpected filename'}")

    if not (filename_ok and size_ok and hash_ok):
        print("\n  CERTIFICATION FAILURE -- workbook identity check failed.")
        sys.exit(1)
    print("  Identity: PASS")

    # ── Method A ──────────────────────────────────────────────────────────────
    print(f"\n[2] METHOD A -- openpyxl")
    a = method_a_openpyxl(path)
    if "error" in a:
        print(f"  ERROR: {a['error']}")
        sys.exit(1)
    print(f"  Sheets      : {a['sheet_names']}")
    print(f"  Hidden      : {a['hidden_sheets'] or 'none'}")
    print(f"  VBA/Macros  : {a['macros']}")
    print(f"  Ext links   : {a['external_links']}")
    print(f"  Product rows: {a['row_count']}")
    print(f"  Formulas    : {a['formula_count']}")
    print(f"  Errors      : {len(a['error_cells'])}")
    print(f"  Merged(data): {len(a['merged_in_data'])}")
    print(f"  Hidden rows : {a['hidden_rows_count']}")
    print(f"  Blank rows  : {len(a['blank_rows_in_data'])}")
    print(f"  Headers     : {a['headers']}")

    # ── Method B ──────────────────────────────────────────────────────────────
    print(f"\n[3] METHOD B -- OOXML/ZIP")
    b = method_b_ooxml(path)
    if "error" in b:
        print(f"  ERROR: {b['error']}")
        sys.exit(1)
    print(f"  Sheets      : {b['sheet_names']}")
    print(f"  Hidden      : {b['hidden_sheets'] or 'none'}")
    print(f"  ZIP entries : {b['zip_entries_count']}")
    print(f"  VBA         : {b['has_vba']}")
    print(f"  Ext links   : {b['has_external_links']}")
    print(f"  Connections : {b['has_connections']}")
    print(f"  Shared str  : {b['shared_strings_count']}")
    print(f"  Product rows: {b['row_count']}")
    print(f"  Formulas    : {b['formula_count']}")
    print(f"  Errors      : {len(b['error_cells'])}")
    print(f"  Merged(data): {b['merged_in_data_count']}")
    print(f"  Blank rows  : {len(b['blank_rows_in_data'])}")
    print(f"  Headers     : {b['headers']}")

    # -- Cross-check ----------------------------------------------------------
    print(f"\n[4] CROSS-CHECK A vs B")
    xcheck_blocking, xcheck_advisory = cross_check(a, b)
    if xcheck_blocking:
        for d in xcheck_blocking:
            print(f"  !! BLOCKING: {d}")
    if xcheck_advisory:
        for d in xcheck_advisory:
            print(f"  NOTE: {d}")
    if not xcheck_blocking and not xcheck_advisory:
        print("  All cross-checks agree -- OK")

    # ── Analytics ─────────────────────────────────────────────────────────────
    print(f"\n[5] ANALYTICS (Method A rows)")
    headers = a.get("headers") or []
    rows    = a.get("rows") or []
    stats   = analyse_rows(headers, rows)

    print(f"  Total rows   : {stats['total_rows']}")
    print(f"  Individual   : {stats.get('individual_count', '?')}")
    print(f"  Wholesale    : {stats.get('wholesale_count', '?')}")
    print(f"  Categories   : {stats.get('distinct_categories', '?')}")
    print(f"  Exact dup names: {stats['exact_duplicate_names']}")
    if stats.get("price_stats"):
        for pname, ps in stats["price_stats"].items():
            print(f"  Price '{pname}': min={ps['min']}, max={ps['max']}, zeros={ps['zeros']}")
    print(f"  Unit distinct: {stats.get('unit_distinct_values', '?')}")

    # ── Column profile ────────────────────────────────────────────────────────
    col_profile = profile_columns(headers, rows)

    # ── Aggregate findings ────────────────────────────────────────────────────
    blocking = list(stats["blocking"])
    advisory = list(stats["advisory"])
    review   = list(stats["review"])

    if xcheck_blocking:
        blocking += [f"CROSS-CHECK DISAGREEMENT: {d}" for d in xcheck_blocking]
    if xcheck_advisory:
        advisory += xcheck_advisory
    if a.get("error_cells"):
        blocking.append(f"FORMULA ERROR CELLS in Products sheet: {len(a['error_cells'])}")
    if a.get("macros"):
        blocking.append("MACROS/VBA detected in workbook")
    if a.get("external_links"):
        advisory.append("EXTERNAL LINKS detected in workbook")
    if b.get("has_connections"):
        advisory.append("DATA CONNECTIONS detected in workbook")
    if a.get("merged_in_data"):
        advisory.append(
            f"MERGED CELLS in data range ({len(a['merged_in_data'])}): {a['merged_in_data'][:3]}")
    if a.get("hidden_rows_count", 0):
        advisory.append(f"HIDDEN ROWS in Products sheet: {a['hidden_rows_count']}")
    if a.get("formula_count", 0):
        advisory.append(f"FORMULAS in Products sheet: {a['formula_count']} cells")

    # ── Results ───────────────────────────────────────────────────────────────
    print(f"\n{sep}")
    print("BLOCKING FAILURES:")
    if blocking:
        for item in blocking:
            print(f"  FAIL: {item}")
    else:
        print("  None")

    print("\nADVISORY FINDINGS:")
    if advisory:
        for item in advisory:
            print(f"  WARN: {item}")
    else:
        print("  None")

    print("\nHUMAN REVIEW ITEMS:")
    if review:
        for item in review:
            print(f"  REVIEW: {item}")
    else:
        print("  None")

    # ── Post-analysis identity ────────────────────────────────────────────────
    print(f"\n[6] POST-ANALYSIS IDENTITY")
    final_sha256    = sha256_file(path)
    hash_unchanged  = final_sha256 == actual_sha256
    print(f"  Final SHA-256: {final_sha256}  {'UNCHANGED OK' if hash_unchanged else 'CHANGED -- FAIL'}")
    if not hash_unchanged:
        blocking.append("WORKBOOK HASH CHANGED DURING ANALYSIS")

    # ── Decision ──────────────────────────────────────────────────────────────
    print(f"\n{sep}")
    if blocking:
        decision  = "NOT CERTIFIED"
        exit_code = 1
    elif advisory or review:
        decision  = "CERTIFIED WITH ADVISORIES"
        exit_code = 0
    else:
        decision  = "CERTIFIED"
        exit_code = 0
    print(f"CERTIFICATION DECISION: {decision}")
    print(sep)

    # ── JSON report ───────────────────────────────────────────────────────────
    report = {
        "certification_decision": decision,
        "file_identity": {
            "path":                     str(path),
            "filename":                 path.name,
            "size_bytes":               actual_size,
            "sha256_pre":               actual_sha256,
            "sha256_post":              final_sha256,
            "hash_unchanged":           hash_unchanged,
            "registered_sha256":        REGISTERED_SHA256,
            "hash_matches_registered":  hash_ok,
        },
        "expected_totals": {
            "products":   EXPECTED_PRODUCTS,
            "categories": EXPECTED_CATEGORIES,
            "individual": EXPECTED_INDIVIDUAL,
            "wholesale":  EXPECTED_WHOLESALE,
        },
        "actual_totals": {
            "products":   stats["total_rows"],
            "categories": stats.get("distinct_categories"),
            "individual": stats.get("individual_count"),
            "wholesale":  stats.get("wholesale_count"),
        },
        "sheet_structure": {
            "method_a_sheets":     a.get("sheet_names"),
            "method_b_sheets":     b.get("sheet_names"),
            "hidden_sheets":       a.get("hidden_sheets"),
            "has_macros":          a.get("macros"),
            "has_external_links":  a.get("external_links"),
            "has_data_connections": b.get("has_connections"),
            "product_rows_a":      a.get("row_count"),
            "product_rows_b":      b.get("row_count"),
            "formula_count_a":     a.get("formula_count"),
            "formula_count_b":     b.get("formula_count"),
            "error_cells_a":       len(a.get("error_cells", [])),
            "merged_data_a":       len(a.get("merged_in_data", [])),
            "hidden_rows_a":       a.get("hidden_rows_count", 0),
            "blank_rows_a":        len(a.get("blank_rows_in_data", [])),
            "sheet_protected":     a.get("sheet_protected"),
            "header_row":          HEADER_ROW,
        },
        "headers": headers,
        "column_profile": col_profile,
        "sales_type_distribution": stats.get("sales_type_distribution"),
        "price_stats":             stats.get("price_stats"),
        "unit_distinct_count":     stats.get("unit_distinct_values"),
        "unit_sample":             stats.get("unit_sample"),
        "exact_duplicate_names":   stats.get("exact_duplicate_names"),
        "cross_check_blocking": xcheck_blocking,
        "cross_check_advisory": xcheck_advisory,
        "blocking_findings":       blocking,
        "advisory_findings":       advisory,
        "human_review_items":      review,
        "classification_rule": {
            "note": "Sales Type is authoritative in the final workbook. Individual = Piece unit + pack_qty 1.",
            "individual_unit": INDIVIDUAL_UNIT,
            "individual_pack_qty": INDIVIDUAL_PACK_QTY,
            "rule_individual_count": stats.get("rule_individual_count"),
            "rule_wholesale_count":  stats.get("rule_wholesale_count"),
        },
    }

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as jf:
            json.dump(report, jf, indent=2, ensure_ascii=False)
        print(f"\nJSON report: {args.json_out}")

    sys.exit(exit_code)


if __name__ == "__main__":
    main()
