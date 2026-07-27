"""
RS-Database-Updated.xlsx Generator
Merges WS RATES.pdf data with catalogue-products.csv
"""

import json, re, csv, math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers, Border, Side
from openpyxl.utils import get_column_letter

# ── 1. Parse PDF raw lines ──────────────────────────────────────────────
with open('.codex-phase7-tmp/pdf_raw_lines.json', 'r', encoding='utf-8') as f:
    pdf_lines = json.load(f)

CAT_PATTERN = r'[A-Za-z][A-Za-z &/+/\-]*'
MAIN_RE = re.compile(r'(?:^|\s)(\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s*(' + CAT_PATTERN + r')$')

pdf_records = []
for line in pdf_lines:
    line = line.strip()
    if not line:
        continue
    m = MAIN_RE.search(line)
    if m:
        ws = float(m.group(1))
        cost = float(m.group(2))
        cat_code = m.group(3).strip()
        name = line[:m.start()].strip()
        if name and name[-1] in ' \t':
            name = name[:-1]
        pdf_records.append({
            'name': name,
            'wholesale': ws,
            'cost': cost,
            'cat_code': cat_code,
        })

print(f"PDF lines parsed: {len(pdf_records)}")

# ── 2. Read existing CSV ────────────────────────────────────────────────
csv_rows = []
with open('.codex-phase7-tmp/catalogue-products.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        csv_rows.append({
            'item_name': row['Item Name'],
            'category': row['Category'],
            'sales_type': row['Sales Type'],
            'wholesale': float(row['Wholesale Price']) if row['Wholesale Price'] else 0.0,
        })

print(f"CSV rows: {len(csv_rows)}")

# ── 3. Category code → category name mapping ───────────────────────────
CATEGORY_MAP = {
    "BALL POINT": "Ballpoint Pens",
    "PENCILS": "Graphite & Mechanical Pencils",
    "PENCIL COLOR": "Colored Pencils",
    "PENCIL JAR": "Graphite & Mechanical Pencils",
    "FANCY PENCIL": "Graphite & Mechanical Pencils",
    "ERASERS": "Erasers",
    "SHARPNERS": "Sharpeners",
    "STAPLERS": "Staplers",
    "STAPLER PIN": "Staples, Pins & Tacks",
    "TAPE": "Adhesive Tapes & Dispensers",
    "SCOTCH TAPE": "Adhesive Tapes & Dispensers",
    "GUM STICK": "Adhesives & Glue",
    "GERMAN GLUE": "Adhesives & Glue",
    "UHU": "Adhesives & Glue",
    "GLUE GUN": "Adhesives & Glue",
    "GLUE GUN ROD": "Adhesives & Glue",
    "REGISTERS": "Registers",
    "GATTA REGISTER": "Account Books & Ledgers",
    "RING REGISTERS": "Registers",
    "DIARIES": "Diaries & Homework Diaries",
    "FANCY DIARIES": "Diaries & Homework Diaries",
    "copies": "Exercise Books & School Copies",
    "drawing copy": "Practical, Graph & Drawing Books",
    "PRACTICAL COPY": "Practical, Graph & Drawing Books",
    "GRAPH COPY": "Practical, Graph & Drawing Books",
    "CASH MEMO": "Receipt, Attendance & Cash Memo Books",
    "SIPARA": "Religious Books & Learning",
    "QURAN-E-PAK": "Religious Books & Learning",
    "QAIDA": "Religious Books & Learning",
    "SIPARA SET": "Religious Books & Learning",
    "CHINA": "Novelty Toys & Collectible Games",
    "CHINA DAIRIES": "Novelty Toys & Collectible Games",
    "BIRTHDAY": "Party Decorations & Balloons",
    "BALLOONS": "Party Decorations & Balloons",
    "DISPLAY": "Party Decorations & Balloons",
    "CANDLES": "Candles, Poppers & Celebration Supplies",
    "BAGS": "Bags, Pouches & Cases",
    "POUCH": "Bags, Pouches & Cases",
    "LIFAFA": "Envelopes",
    "FILE": "Files, Folders & Document Storage",
    "CARD COVER": "Covers & Protective Sheets",
    "PLASTIC SHEET/ROLL": "Packaging & Storage Materials",
    "CLIP BOARD": "Desk & Office Accessories",
    "INK PEN": "Fountain & Ink Pens",
    "Fountain pen": "Fountain & Ink Pens",
    "GEL PEN": "Gel Pens",
    "POINTER": "Fineliners, Pointers & Signature Pens",
    "BOARD & PERMANENT": "Markers & Highlighters",
    "MARKER": "Markers & Highlighters",
    "CUT MARKERS": "Markers & Highlighters",
    "MARKER COLOR": "Markers & Highlighters",
    "PAINT MARKERS": "Markers & Highlighters",
    "HIGHLIGHTER": "Markers & Highlighters",
    "BOARD INK": "Inks & Refills",
    "FOUNTAIN INK": "Inks & Refills",
    "Correction pen": "Correction Supplies",
    "REMOVERS": "Correction Supplies",
    "GEOMATRIES": "Geometry & Mathematical Instruments",
    "FANCY GEOMETRY": "Geometry & Mathematical Instruments",
    "D-SET SQARE": "Geometry & Mathematical Instruments",
    "SCALES": "Rulers & Scales",
    "COMPASS": "Geometry & Mathematical Instruments",
    "BAT": "Cricket Equipment",
    "BAT GRIP": "Cricket Equipment",
    "BALL": "Balls & Outdoor Play",
    "RACKETS": "Badminton & Racket Sports",
    "SCIENTIFIC CALCULAT": "Calculators",
    "SHOP CALCULATOR": "Calculators",
    "PAPER": "Paper - Other",
    "PAPER WRITE": "Printing, Photocopy & Photo Paper",
    "COLOUR PAPER": "Craft, Chart & Specialty Paper",
    "COMPUTER PAPER WH": "Printing, Photocopy & Photo Paper",
    "CHARTS": "Craft, Chart & Specialty Paper",
    "GIFT SHEET": "Gift Wrapping & Ribbons",
    "STICKERS": "Labels, Tags & Stickers",
    "STICKY NOTE": "Sticky Notes & Memo Pads",
    "WHITE BOARD": "Whiteboards, Slates & Board Accessories",
    "SLATES": "Whiteboards, Slates & Board Accessories",
    "LUDO": "Board & Indoor Games",
    "LUDO DANA + GOTI": "Board & Indoor Games",
    "WATER COLORS": "Paints & Color Media",
    "POSTER COLORS": "Paints & Color Media",
    "PAINT BRUSH": "Brushes & Painting Tools",
    "BRUSHES": "Brushes & Painting Tools",
    "CANVAS BOARD": "Canvas, Drawing & Sketching Supplies",
    "CRAYONS": "Crayons & Oil Pastels",
    "COLORING KIT": "Art Materials - Other",
    "GLITTER SHEETS": "Glitter, Foam & Craft Materials",
    "GLITTER DUST": "Glitter, Foam & Craft Materials",
    "CLAY": "Clay, Dough & Modeling Materials",
    "STICKS": "Glitter, Foam & Craft Materials",
    "SCISSOR": "Scissors",
    "CUTTERS": "Cutters, Blades & Cutting Tools",
    "KEY CHAINS": "Keychains & Small Accessories",
    "COUNTER": "Tally Counters",
    "THERMAL ROLL": "Paper - Other",
    "CARTRIDGE": "Inks & Refills",
    "NEW LIFE": "General Pens & Pen Sets",
    "HAND GRIP": "Sports & Fitness Accessories",
    "D": "Writing Instruments - Model Unclear",
    "LR": "Writing Instruments - Model Unclear",
    "STEAKER": "Labels, Tags & Stickers",
    "BALLOONS": "Party Decorations & Balloons",
    "SCOTCH TAPE": "Adhesive Tapes & Dispensers",
}

# Fill any missing codes with a cleaned-up title-case version
CSV_CATEGORIES = set(r['category'] for r in csv_rows)
for code in list(CATEGORY_MAP.keys()):
    if CATEGORY_MAP[code] not in CSV_CATEGORIES:
        # Try to find the closest match
        pass

# Build reverse lookup: for each code, find best matching CSV category
def normalize_cat_name(name):
    return re.sub(r'\s+', ' ', name.strip().lower())

cat_map_clean = {normalize_cat_name(v): v for v in CSV_CATEGORIES}

# Ensure all codes map to valid CSV categories
final_cat_map = {}
for code, mapped in CATEGORY_MAP.items():
    clean_mapped = normalize_cat_name(mapped)
    if clean_mapped in cat_map_clean:
        final_cat_map[code] = cat_map_clean[clean_mapped]
    else:
        # Try to find the best match
        best = None
        for cname in CSV_CATEGORIES:
            if code.lower() in cname.lower() or cname.lower() in code.lower():
                best = cname
                break
        if best:
            final_cat_map[code] = best
        else:
            final_cat_map[code] = mapped

# For codes not in map, generate from CSV categories
mapped_codes = set(final_cat_map.keys())
unmapped_codes = set(r['cat_code'] for r in pdf_records) - mapped_codes
for code in unmapped_codes:
    # Try fuzzy match against all CSV categories
    cl = code.lower()
    best = None
    best_score = 0
    for cname in CSV_CATEGORIES:
        cnl = cname.lower()
        # Simple word-overlap scoring
        code_words = set(cl.split())
        cname_words = set(cnl.split())
        overlap = len(code_words & cname_words)
        if overlap > best_score:
            best_score = overlap
            best = cname
    if best and best_score > 0:
        final_cat_map[code] = best
    else:
        final_cat_map[code] = code.title()

print(f"Category codes mapped: {len(final_cat_map)}")

# ── 4. Name normalization for matching ──────────────────────────────────
def normalize_name(name):
    """Normalize product name for matching."""
    n = name.lower().strip()
    # Remove special chars that differ between sources
    n = re.sub(r'[`\'\"\(\)\[\]]', '', n)
    # Normalize whitespace
    n = re.sub(r'\s+', ' ', n)
    # Remove trailing backticks, commas, periods
    n = n.rstrip('`.,;:')
    return n.strip()

def name_match_key(name):
    """Generate a matching key."""
    n = normalize_name(name)
    return n

# Build PDF index by normalized name
pdf_by_norm = defaultdict(list)
for rec in pdf_records:
    key = name_match_key(rec['name'])
    pdf_by_norm[key].append(rec)

# Also build index by name with price as tiebreaker
def name_with_price_key(name, price):
    n = normalize_name(name)
    return (n, round(price, 1))

# ── 5. Match products ───────────────────────────────────────────────────
matched = []
csv_matched = set()  # track CSV indices that were matched
stats = {
    'matched': 0,
    'csv_only': 0,
    'pdf_extra': 0,
    'zero_neg_cost': 0,
    'negative_profit': 0,
    'category_changed': 0,
    'price_mismatch': 0,
}

pdf_used = set()

for csv_idx, csv_row in enumerate(csv_rows):
    csv_name = csv_row['item_name']
    csv_ws = csv_row['wholesale']
    key = name_match_key(csv_name)
    candidates = pdf_by_norm.get(key, [])

    if not candidates:
        continue  # will be added as CSV-only later

    # Pick the best candidate: prefer by exact name match, then by price proximity
    best = None
    best_diff = float('inf')
    for c in candidates:
        if id(c) in pdf_used:
            continue
        # Compare full normalized names (not just key) for better matching
        diff = abs(c['wholesale'] - csv_ws)
        if diff < best_diff:
            best_diff = diff
            best = c

    if best is None:
        continue  # all candidates already used

    pdf_used.add(id(best))
    csv_matched.add(csv_idx)

    cost_price = best['cost']
    has_neg_cost = cost_price < 0
    if has_neg_cost:
        stats['zero_neg_cost'] += 1

    profit = (csv_ws - cost_price) if (csv_ws > 0 and cost_price >= 0) else None
    margin = (profit / csv_ws * 100) if (profit is not None and csv_ws > 0) else None

    if profit is not None and profit < 0:
        stats['negative_profit'] += 1

    pdf_cat = final_cat_map.get(best['cat_code'], best['cat_code'])
    cat_changed = pdf_cat != csv_row['category']
    if cat_changed:
        stats['category_changed'] += 1

    ws_mismatch = abs(best['wholesale'] - csv_ws) > 0.01
    if ws_mismatch:
        stats['price_mismatch'] += 1

    flags = []
    if has_neg_cost:
        flags.append(f"Negative cost ({cost_price})")
    if ws_mismatch:
        flags.append(f"WS price diff: CSV={csv_ws:.2f} PDF={best['wholesale']:.2f}")
    if cat_changed:
        flags.append(f"Category: {csv_row['category']} \u2192 {pdf_cat}")
    if profit is not None and profit < 0:
        flags.append("Negative profit")
    if cost_price == 0:
        flags.append("Zero cost price")

    matched.append({
        'csv_name': csv_name,
        'category': csv_row['category'],
        'sales_type': csv_row['sales_type'],
        'csv_ws': csv_ws,
        'cost_price': cost_price,
        'updated_category': pdf_cat if cat_changed else '',
        'profit': profit,
        'margin': margin,
        'match_status': 'Matched',
        'flags': '; '.join(flags) if flags else '',
        'pdf_ws': best['wholesale'],
    })
    stats['matched'] += 1

# CSV-only: rows that were never matched
csv_rows_unmatched = [csv_rows[i] for i in range(len(csv_rows)) if i not in csv_matched]
for csv_row in csv_rows_unmatched:
    matched.append({
        'csv_name': csv_row['item_name'],
        'category': csv_row['category'],
        'sales_type': csv_row['sales_type'],
        'csv_ws': csv_row['wholesale'],
        'cost_price': 'Not found in PDF',
        'updated_category': '',
        'profit': '',
        'margin': '',
        'match_status': 'Not in PDF',
        'flags': '',
        'pdf_ws': '',
    })
    stats['csv_only'] += 1

# PDF-extra: PDF records never matched
all_csv_norm = set()
for i in range(len(csv_rows)):
    all_csv_norm.add((name_match_key(csv_rows[i]['item_name']), round(csv_rows[i]['wholesale'], 1)))

for rec in pdf_records:
    if id(rec) in pdf_used:
        continue
    # Check if this exact name+price combo exists in CSV unmatched (avoid adding duplicates)
    rec_key = (name_match_key(rec['name']), round(rec['wholesale'], 1))
    if rec_key not in all_csv_norm:
        pdf_cat = final_cat_map.get(rec['cat_code'], rec['cat_code'])
        cost_price = rec['cost']
        has_neg_cost = cost_price < 0
        if has_neg_cost:
            stats['zero_neg_cost'] += 1

        flags = []
        if has_neg_cost:
            flags.append(f"Negative cost ({cost_price})")
        if cost_price == 0:
            flags.append("Zero cost price")

        matched.append({
            'csv_name': rec['name'],
            'category': pdf_cat,
            'sales_type': '',
            'csv_ws': rec['wholesale'],
            'cost_price': cost_price,
            'updated_category': '',
            'profit': '',
            'margin': '',
            'match_status': 'PDF Extra',
            'flags': '; '.join(flags) if flags else '',
            'pdf_ws': rec['wholesale'],
        })
        stats['pdf_extra'] += 1

# Count zero/negative cost prices
zero_neg_count = sum(1 for m in matched if isinstance(m['cost_price'], (int, float)) and m['cost_price'] <= 0)

# ── 6. Create Excel workbook ────────────────────────────────────────────
wb = Workbook()

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_alignment = Alignment(vertical='center')
price_format = '#,##0.00'
currency_format = '#,##0.00'
pct_format = '0.00%'
thin_border = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3'),
)

def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def auto_width(ws, max_width=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)

# ── Sheet 1: Summary ────────────────────────────────────────────────────
ws_summary = wb.active
ws_summary.title = "Summary"

total_csv = len(csv_rows)
total_pdf = len(pdf_records)
total_combined = len(matched)

summary_data = [
    ("RS-Database-Updated.xlsx — Merge Report", ""),
    ("", ""),
    ("Source Data", ""),
    ("CSV products (catalogue-products.csv)", total_csv),
    ("PDF products (WS RATES.pdf)", total_pdf),
    ("Total combined rows", total_combined),
    ("", ""),
    ("Match Results", ""),
    ("Matched products", stats['matched']),
    ("CSV only (not in PDF)", stats['csv_only']),
    ("PDF extra (not in CSV)", stats['pdf_extra']),
    ("", ""),
    ("Data Quality", ""),
    ("Zero or negative cost prices", zero_neg_count),
    ("Wholesale price mismatches", stats['price_mismatch']),
    ("Category changes", stats['category_changed']),
    ("Negative profit entries", stats['negative_profit']),
    ("", ""),
    ("Products with Negative / Erroneous Cost Prices", ""),
]

# List zero/negative cost entries
neg_cost_list = [m for m in matched if isinstance(m['cost_price'], (int, float)) and m['cost_price'] <= 0]
for m in neg_cost_list:
    summary_data.append((m['csv_name'], m['cost_price']))

ws_summary.cell(row=1, column=1, value="RS-Database-Updated.xlsx — Merge Report")
ws_summary.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=14, color='2F5496')

for i, (label, value) in enumerate(summary_data, start=3):
    ws_summary.cell(row=i, column=1, value=label)
    if value != '':
        ws_summary.cell(row=i, column=2, value=value)
    if label in ("Source Data", "Match Results", "Data Quality", "Products with Negative / Erroneous Cost Prices"):
        ws_summary.cell(row=i, column=1).font = Font(name='Calibri', bold=True, size=11, color='2F5496')

ws_summary.column_dimensions['A'].width = 50
ws_summary.column_dimensions['B'].width = 20

# ── Sheet 2: Products (main data) ───────────────────────────────────────
ws_products = wb.create_sheet("Products")

headers = [
    "Row ID", "Item Name", "Category", "Sales Type", "Wholesale Price",
    "Cost Price / Buying Rate", "Updated Category", "Profit", "Profit Margin %",
    "Match Status", "Flags"
]
for col, header in enumerate(headers, 1):
    ws_products.cell(row=1, column=col, value=header)
style_header(ws_products, len(headers))

for i, m in enumerate(matched, start=2):
    ws_products.cell(row=i, column=1, value=i - 1)
    ws_products.cell(row=i, column=2, value=m['csv_name'])
    ws_products.cell(row=i, column=3, value=m['category'])
    ws_products.cell(row=i, column=4, value=m['sales_type'])
    ws_products.cell(row=i, column=5, value=m['csv_ws'])

    cp = m['cost_price']
    if isinstance(cp, (int, float)):
        ws_products.cell(row=i, column=6, value=round(cp, 2))
        ws_products.cell(row=i, column=6).number_format = price_format
    else:
        ws_products.cell(row=i, column=6, value=str(cp))

    ws_products.cell(row=i, column=7, value=m['updated_category'])

    profit = m['profit']
    if isinstance(profit, (int, float)):
        ws_products.cell(row=i, column=8, value=round(profit, 2))
        ws_products.cell(row=i, column=8).number_format = price_format
    else:
        ws_products.cell(row=i, column=8, value='')

    margin = m['margin']
    if isinstance(margin, (int, float)):
        ws_products.cell(row=i, column=9, value=round(margin, 2))
        ws_products.cell(row=i, column=9).number_format = '0.00'
    else:
        ws_products.cell(row=i, column=9, value='')

    ws_products.cell(row=i, column=10, value=m['match_status'])
    ws_products.cell(row=i, column=11, value=m['flags'])

    # Color-code match status
    status_cell = ws_products.cell(row=i, column=10)
    if m['match_status'] == 'Matched':
        status_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    elif m['match_status'] == 'Not in PDF':
        status_cell.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    elif m['match_status'] == 'PDF Extra':
        status_cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    # Flag rows with issues
    if m['flags']:
        for col in range(1, len(headers) + 1):
            ws_products.cell(row=i, column=col).fill = PatternFill(
                start_color='FFF9C4', end_color='FFF9C4', fill_type='solid'
            )

auto_width(ws_products)
# Freeze top row
ws_products.freeze_panes = 'A2'

# ── Sheet 3: Mismatches ────────────────────────────────────────────────
ws_mismatch = wb.create_sheet("Mismatches")

mismatch_headers = [
    "Row ID", "Item Name", "CSV Category", "PDF Category",
    "CSV Wholesale", "PDF Wholesale", "Cost Price", "Match Status", "Flags"
]
for col, header in enumerate(mismatch_headers, 1):
    ws_mismatch.cell(row=1, column=col, value=header)
style_header(ws_mismatch, len(mismatch_headers))

mismatch_rows = [
    m for m in matched
    if m['flags'] or m['updated_category']
]
row_num = 2
for m in mismatch_rows:
    ws_mismatch.cell(row=row_num, column=1, value=row_num - 1)
    ws_mismatch.cell(row=row_num, column=2, value=m['csv_name'])
    ws_mismatch.cell(row=row_num, column=3, value=m['category'])
    ws_mismatch.cell(row=row_num, column=4, value=m['updated_category'] if m['updated_category'] else m['category'])
    ws_mismatch.cell(row=row_num, column=5, value=m['csv_ws'])
    ws_mismatch.cell(row=row_num, column=6, value=m['pdf_ws'] if isinstance(m['pdf_ws'], (int, float)) else '')

    cp = m['cost_price']
    if isinstance(cp, (int, float)):
        ws_mismatch.cell(row=row_num, column=7, value=round(cp, 2))
        ws_mismatch.cell(row=row_num, column=7).number_format = price_format
    else:
        ws_mismatch.cell(row=row_num, column=7, value=str(cp))

    ws_mismatch.cell(row=row_num, column=8, value=m['match_status'])
    ws_mismatch.cell(row=row_num, column=9, value=m['flags'])
    row_num += 1

auto_width(ws_mismatch)
ws_mismatch.freeze_panes = 'A2'

# ── Sheet 4: Categories ────────────────────────────────────────────────
ws_cats = wb.create_sheet("Categories")

cat_headers = [
    "Category Name", "Product Count", "Matched", "Not in PDF", "PDF Extra",
    "Avg Wholesale", "Avg Cost", "Total Wholesale", "Avg Profit", "Avg Margin %"
]
for col, header in enumerate(cat_headers, 1):
    ws_cats.cell(row=1, column=col, value=header)
style_header(ws_cats, len(cat_headers))

# Aggregate by category
cat_agg = defaultdict(lambda: {
    'count': 0, 'matched': 0, 'not_in_pdf': 0, 'pdf_extra': 0,
    'ws_sum': 0, 'cost_sum': 0, 'cost_count': 0, 'profit_sum': 0, 'profit_count': 0,
    'margin_sum': 0, 'margin_count': 0,
})

for m in matched:
    cat = m['category']
    agg = cat_agg[cat]
    agg['count'] += 1
    if m['match_status'] == 'Matched':
        agg['matched'] += 1
    elif m['match_status'] == 'Not in PDF':
        agg['not_in_pdf'] += 1
    elif m['match_status'] == 'PDF Extra':
        agg['pdf_extra'] += 1

    ws_val = m['csv_ws'] if isinstance(m['csv_ws'], (int, float)) else 0
    agg['ws_sum'] += ws_val

    cp = m['cost_price']
    if isinstance(cp, (int, float)) and cp > 0:
        agg['cost_sum'] += cp
        agg['cost_count'] += 1

    profit = m['profit']
    if isinstance(profit, (int, float)):
        agg['profit_sum'] += profit
        agg['profit_count'] += 1

    margin = m['margin']
    if isinstance(margin, (int, float)):
        agg['margin_sum'] += margin
        agg['margin_count'] += 1

row_num = 2
for cat_name in sorted(cat_agg.keys()):
    agg = cat_agg[cat_name]
    ws_cats.cell(row=row_num, column=1, value=cat_name)
    ws_cats.cell(row=row_num, column=2, value=agg['count'])
    ws_cats.cell(row=row_num, column=3, value=agg['matched'])
    ws_cats.cell(row=row_num, column=4, value=agg['not_in_pdf'])
    ws_cats.cell(row=row_num, column=5, value=agg['pdf_extra'])
    ws_cats.cell(row=row_num, column=6, value=round(agg['ws_sum'] / agg['count'], 2) if agg['count'] > 0 else 0)
    ws_cats.cell(row=row_num, column=7, value=round(agg['cost_sum'] / agg['cost_count'], 2) if agg['cost_count'] > 0 else '')
    ws_cats.cell(row=row_num, column=8, value=round(agg['ws_sum'], 2))
    ws_cats.cell(row=row_num, column=9, value=round(agg['profit_sum'] / agg['profit_count'], 2) if agg['profit_count'] > 0 else '')
    ws_cats.cell(row=row_num, column=10, value=round(agg['margin_sum'] / agg['margin_count'], 2) if agg['margin_count'] > 0 else '')
    row_num += 1

auto_width(ws_cats)
ws_cats.freeze_panes = 'A2'

# ── Sheet 5: Review Queue ──────────────────────────────────────────────
ws_review = wb.create_sheet("Review Queue")

review_headers = [
    "Row ID", "Item Name", "Category", "Issue Type", "Current Value", "PDF Value", "Recommendation"
]
for col, header in enumerate(review_headers, 1):
    ws_review.cell(row=1, column=col, value=header)
style_header(ws_review, len(review_headers))

review_rows = []
for m in matched:
    if not m['flags']:
        continue

    for flag in m['flags'].split('; '):
        if not flag:
            continue
        row_data = [m['csv_name'], m['category']]

        if 'Negative cost' in flag:
            review_rows.append((m['csv_name'], m['category'],
                                'Negative / Erroneous Cost',
                                str(m['cost_price']),
                                'Check source data',
                                'Verify cost price with supplier'))
        elif 'WS price diff' in flag:
            review_rows.append((m['csv_name'], m['category'],
                                'Wholesale Price Mismatch',
                                f"CSV: {m['csv_ws']}",
                                str(m['pdf_ws']) if isinstance(m.get('pdf_ws'), (int, float)) else '',
                                'Verify which price is correct'))
        elif 'Category' in flag and '→' in flag:
            parts = flag.split('→')
            old_cat = parts[0].replace('Category: ', '').strip() if len(parts) > 0 else ''
            new_cat = parts[1].strip() if len(parts) > 1 else ''
            review_rows.append((m['csv_name'], m['category'],
                                'Category Change',
                                old_cat,
                                new_cat,
                                'Confirm category assignment'))
        elif 'Negative profit' in flag:
            review_rows.append((m['csv_name'], m['category'],
                                'Negative Profit',
                                f"Profit: {m['profit']}",
                                'Review pricing strategy',
                                'Consider cost price or wholesale adjustment'))
        elif 'Zero cost' in flag:
            review_rows.append((m['csv_name'], m['category'],
                                'Zero Cost Price',
                                '0.00',
                                'Missing cost data',
                                'Obtain cost price from supplier'))

# Remove duplicates
seen = set()
unique_reviews = []
for row in review_rows:
    key = (row[0], row[2])
    if key not in seen:
        seen.add(key)
        unique_reviews.append(row)

for i, review in enumerate(unique_reviews, start=2):
    ws_review.cell(row=i, column=1, value=i - 1)
    ws_review.cell(row=i, column=2, value=review[0])
    ws_review.cell(row=i, column=3, value=review[1])
    ws_review.cell(row=i, column=4, value=review[2])
    ws_review.cell(row=i, column=5, value=review[3])
    ws_review.cell(row=i, column=6, value=review[4])
    ws_review.cell(row=i, column=7, value=review[5])

auto_width(ws_review)
ws_review.freeze_panes = 'A2'

# ── Fix the review queue flag parsing ──────────────────────────────────
# (Already handled in the loop above - ws_price fix is embedded)

# ── Save ────────────────────────────────────────────────────────────────
output_path = 'RS-Database-Updated.xlsx'
wb.save(output_path)

print(f"\n{'='*60}")
print(f"File saved: {output_path}")
print(f"{'='*60}")
print(f"Total products in spreadsheet: {len(matched)}")
print(f"  Matched: {stats['matched']}")
print(f"  Not in PDF (CSV only): {stats['csv_only']}")
print(f"  PDF Extra: {stats['pdf_extra']}")
print(f"")
print(f"Zero or negative cost prices: {zero_neg_count}")
print(f"Category changes: {stats['category_changed']}")
print(f"Wholesale price mismatches: {stats['price_mismatch']}")
print(f"Negative profit entries: {stats['negative_profit']}")
print(f"Review queue items: {len(unique_reviews)}")
print(f"{'='*60}")
