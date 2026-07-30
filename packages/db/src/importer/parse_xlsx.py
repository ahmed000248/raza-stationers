import sys
import os
import json
import hashlib
from pathlib import Path
import openpyxl

# Exact approved headers list
EXPECTED_HEADERS = [
    'SKU', 'Product Name', 'Category', 'Sales Type', 'Unit of Measure', 
    'Pack Quantity', 'Currency', 'Wholesale Price', 'Buying Price', 
    'Profit', 'Profit Margin %', 'Markup %', 'Active', 'Source Key'
]

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Missing file path argument"}), file=sys.stderr)
        sys.exit(1)
        
    file_path = Path(sys.argv[1])
    if not file_path.exists():
        print(json.dumps({"error": f"File not found: {file_path}"}), file=sys.stderr)
        sys.exit(1)
        
    try:
        # Load workbook in data_only=True mode to read evaluated cached values
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        print(json.dumps({"error": f"Failed to load Excel workbook: {str(e)}"}), file=sys.stderr)
        sys.exit(1)
        
    if "Products" not in wb.sheetnames:
        print(json.dumps({"error": "Worksheet 'Products' not found in workbook"}), file=sys.stderr)
        sys.exit(1)
        
    ws = wb["Products"]
    
    # Read header row (Row 4 in Excel, 1-based index)
    try:
        header_row_vals = [cell.value for cell in list(ws.iter_rows(min_row=4, max_row=4))[0]]
        # Trim trailing None values
        while header_row_vals and header_row_vals[-1] is None:
            header_row_vals.pop()
    except Exception as e:
        print(json.dumps({"error": f"Failed to read header row: {str(e)}"}), file=sys.stderr)
        sys.exit(1)
        
    # Check headers exact match
    if header_row_vals != EXPECTED_HEADERS:
        print(json.dumps({
            "error": "Header validation failed",
            "expected": EXPECTED_HEADERS,
            "found": header_row_vals
        }), file=sys.stderr)
        sys.exit(1)
        
    # Compute header checksum
    header_str = ",".join(EXPECTED_HEADERS)
    header_checksum = hashlib.sha256(header_str.encode('utf-8')).hexdigest()
    
    # Parse rows starting from row 5 (after header)
    rows = []
    row_num = 4
    for row in ws.iter_rows(min_row=5):
        row_num += 1
        cells = [c.value for c in row[:len(EXPECTED_HEADERS)]]
        
        # Skip empty rows
        if all(v is None or str(v).strip() == "" for v in cells):
            continue
            
        # Validate that no required cells have missing cached formula values
        # Profit (index 9), Profit Margin % (index 10), Markup % (index 11) are calculated
        # SKU (0), Name (1), Category (2), Sales Type (3), UOM (4), Pack Qty (5), Currency (6), Prices (7,8)
        # If any of these are un-evaluated (e.g. value is None or starts with '='), reject
        for i, val in enumerate(cells):
            if val is None:
                # Some cells are allowed to be None (like buyingPrice if empty, or profit if inactive)
                # But let's check for un-evaluated formulas
                continue
            if isinstance(val, str) and val.strip().startswith('='):
                print(json.dumps({
                    "error": f"Formula cell not evaluated (missing cached value) at row {row_num}, col {i+1}: {val}"
                }), file=sys.stderr)
                sys.exit(1)
                
        # Structure the raw row
        row_data = {
            "sourceRowNumber": row_num,
            "sourceSheet": "Products",
            "sku": str(cells[0]).strip() if cells[0] is not None else "",
            "name": str(cells[1]).strip() if cells[1] is not None else "",
            "category": str(cells[2]).strip() if cells[2] is not None else "",
            "salesType": str(cells[3]).strip() if cells[3] is not None else "",
            "unitOfMeasure": str(cells[4]).strip() if cells[4] is not None else "",
            "packQuantityRaw": str(cells[5]).strip() if cells[5] is not None else "",
            "currency": str(cells[6]).strip() if cells[6] is not None else "",
            "wholesalePriceRaw": str(cells[7]).strip() if cells[7] is not None else "",
            "buyingPriceRaw": str(cells[8]).strip() if cells[8] is not None else "",
            "profitRaw": str(cells[9]).strip() if cells[9] is not None else "",
            "profitMarginRaw": str(cells[10]).strip() if cells[10] is not None else "",
            "markupRaw": str(cells[11]).strip() if cells[11] is not None else "",
            "activeRaw": str(cells[12]).strip() if cells[12] is not None else "",
            "sourceKey": str(cells[13]).strip() if cells[13] is not None else ""
        }
        rows.append(row_data)
        
    # Print results to stdout as JSON
    print(json.dumps({
        "headerChecksum": header_checksum,
        "rows": rows
    }))

if __name__ == "__main__":
    main()
